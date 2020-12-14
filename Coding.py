from PyQt5.QtCore import pyqtSlot
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtGui import QPixmap
from PyQt5 import Qt
from PyQt5.QtWidgets import QMainWindow, QLabel, QGridLayout, QWidget,QTableWidget,QTableWidgetItem,QGraphicsScene,QGraphicsPixmapItem,QFileDialog
from Design import Ui_MainWindow

import os
import cv2
import math
import numpy as np
import matplotlib.pyplot as plt
from xlrd import open_workbook
from openpyxl.reader.excel import load_workbook
from PIL import Image
from skimage import color
from skimage import data, img_as_float,io
from skimage.measure import structural_similarity as _SSIM
from sklearn.metrics import mean_squared_error as MSE

import Database as db

class override_graphicsScene (Qt.QGraphicsScene):
    def __init__(self,parent = None):
        super(override_graphicsScene,self).__init__(parent)

    def mousePressEvent(self, event):
        super(override_graphicsScene, self).mousePressEvent(event)
        print(event.pos())

class MainWindow(QWidget,Ui_MainWindow):

    _file_path="./temp_image.png"
    _file_path_face = "./temp_face.png"
    
    directory_recognition_face = "./recognitions/"
    directory_recognition_classes = "./classess/"
    
    classes_index_file = ""
    classes_index_dir = ""
    
    selection_cascade = "haarcascade_frontalface_default.xml";
    
    def __init__(self):
        QtWidgets.QMainWindow.__init__(self)
        self.setupUi(self)

        self.btn_img_upload.clicked.connect(self.Btn_Image_Load)
        self.btn_face.clicked.connect(self.Btn_Face)
        self.btn_classes_list.clicked.connect(self.Btn_Classes_Load)
        self.bnt_camera_show.clicked.connect(self.Btn_Camera)
        self.btn_camera_capture.clicked.connect(self.Btn_Camera_Capture)
        self.btn_run.clicked.connect(self.Btn_Run)

        self.cmb_classes.activated[str].connect(self.onActivated_Classes)
        self.cmb_cascade.activated[str].connect(self.onActivated_Cascade)
        self.listView_class.cellClicked.connect(self.onSelected)
        
        self.cmb_classes.addItem("1A")
        self.cmb_classes.addItem("2B")
        self.cmb_classes.addItem("3A")
        
        self.cmb_cascade.addItem("haarcascade")
        self.cmb_cascade.addItem("-")
        
        self.classes_index_file="/1A/1A"
        self.classes_index_dir="/1A/"
        self.img_person.setVisible(True)
        self.chk_all.setVisible(False)
        self.chk_SSIM.setChecked(True)
        
        db.new_db()
        self._Insert_Persons()
    
    def _Insert_Persons(self):
        wb = load_workbook(filename = self.directory_recognition_classes+self.classes_index_file+".xlsx")
        sheet = wb.get_active_sheet()
        #cell.internal_value
        list_excel = []
        for i,row in enumerate(sheet.iter_rows()):
            line=[]
            for j,cell in enumerate(row):
                line.append(str(cell.internal_value))
            
            list_excel.append(line)
        
        
        for i, row in enumerate(list_excel):
            classes_file = self.directory_recognition_classes+self.classes_index_dir+"/Faces/"+str(row[0])+".png"
            number = 0
            name = ''
            dep = ''
            print("row--->",row)
            #if(os.path.isfile(classes_file)):
            #face = cv2.cvtColor(cv2.imread(classes_file),cv2.COLOR_BGR2GRAY)
            if(os.path.isfile(classes_file)):
                face = db.read_file(classes_file)
            else:
                face = db.read_file(self.directory_recognition_classes+"none.png")
                print(self.directory_recognition_classes+"none.png")
            number = row[0]
            name = row[1]
            dep = row[2]
            cls= str(self.classes_index_dir)
            
            db.insert(number,name,dep,face,cls)
            
    def _Insert_Faces(self):
        file_list = os.listdir(self.directory_recognition_face)
        for i, file in enumerate(file_list):
            faces_file = self.directory_recognition_face+file
            print(faces_file)
            #if(os.path.isfile(classes_file)):
            #face = cv2.cvtColor(cv2.imread(classes_file),cv2.COLOR_BGR2GRAY)
            if(os.path.isfile(faces_file)):
                face = db.read_file(faces_file)
            else:
                face = db.read_file(self.directory_recognition_classes+"none.png")
            
            db.insert_faces(file,face)
        
    def onActivated_Classes(self, text):
        self.classes_index_file="/"+text+"/"+text
        self.classes_index_dir="/"+text+"/"
    
    def onActivated_Cascade(self, text):
        if(self.cmb_cascade.currentIndex==0):
            self.selection_cascade="haarcascade_frontalface_default.xml"
        else:
            self.selection_cascade="haarcascade_frontalface_default.xml"
            
    
    def onSelected(self,row,column):
        #print("Row %d and Column %d was clicked" % (row, column))
        #item = self.listView_class.itemAt(row, column)
        #print(item.text())
        id = self.listView_class.item(row,0).text()
        scene = self.show_image(r""+ self.directory_recognition_classes+self.classes_index_dir+"/Faces/"+id+".png")
       
        self.img_person.setScene(scene)
        #print(scene.sceneRect())
        self.img_person.fitInView(scene.sceneRect(),QtCore.Qt.IgnoreAspectRatio)
        #self.img_person.scale(self.img_person.width()/scene.width(),self.img_person.height()/scene.height())
        #self.img_person.setSceneRect(0,0,self.img_person.width(),self.img_person.height())
        
    
    face_count = 0
    
    def Btn_Face(self):
        if(os.path.exists(self._file_path)):
            print("Algılama yapılıyor...")
            
            db.delete_faces()
            
            self.lbl_toplam_kayit.setText("0")
            self.lbl_tanimli.setText("0")
            self.lbl_tanimsiz.setText("0")
            
            self.listView_class.clear()
            self.listView_result.clear()
            
            self.face_count = 0
            
            face_cascade = cv2.CascadeClassifier(self.selection_cascade)
            image = cv2.imread(self._file_path)
            gray = cv2.cvtColor(image,cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray, 1.2, 5)
            
            for i,(x, y, w, h) in enumerate(faces):
                face_img = image[y:y+h, x:x+w]
                face_file_name = self.directory_recognition_face + "face_" + str(i) + ".png"
                
                face_img = cv2.resize(face_img,(75,75))
                cv2.imwrite(face_file_name, face_img)
                cv2.rectangle(image, (x, y), (x+w, y+h), (0, 255, 0), 2)
                
                self.face_count+=1
    
            #image = self.CLAHE(image)
            cv2.imwrite(self._file_path_face, image)   
            scene = self.show_image(r""+self._file_path_face)
            self.img_viewer.setScene(scene)
            
            self._Insert_Faces()
            
            if(self.face_count == 0):
                self.showdialog("Warning","Yüz Algılanamadı.","Belirtilen görüntü içerisinde yüz algılanamadı.")
        else:
            self.showdialog("Warning","Görüntü yüklenemedi.","Lütfen geçerli bir görüntü yükleyiniz.")
    
    score_list=[]
    def Btn_Run(self):
        if(self.face_count != 0):
            print("Karşılaştırılıyor...")
            
            list_persons = db.select()
            list_persons_data = []
            
            list_faces = db.select_faces()
            count=0
                        
            toplam,tanimli,tanimsiz=0,0,0
            
            for i,row in enumerate(list_persons):
                temp_list_person_data = []
                best_score = 0
                toplam+=1
                for j,value in enumerate(row):
                    if(j!=3):
                        temp_list_person_data.append(value)
                
                #print(list_persons[i][3])
                count += 1
                durum = "YOK"
                img_face = db.get_Image(list_persons[i][3],count)
                #img_face = self.rgb2gray(np.asarray(img_face, dtype="int32")) #color.rgb2gray(img_face)

                #file_list = os.listdir(self.directory_recognition_face)
                for index in range(len(list_faces)):
                    #yüzler gray yapılıyor.
                    #img_rec = cv2.imread(self.directory_recognition_face+file_rec)#cv2.cvtColor(,cv2.COLOR_BGR2GRAY)
                    #img_rec = Image.open(self.directory_recognition_face+file_rec)
                    #print("file byte--------->",file_rec)
                    img_rec = db.get_Image(list_faces[index][1],count)
                    #img_rec = self.rgb2gray(np.asarray(img_rec, dtype="int32"))
                    
                    score = self.get_SCORE(img_face,img_rec)
                    print("#####score:",str(score))
                    if(score > best_score):
                        best_score = score
                        durum = "VAR"
                        if(best_score > 0.5):
                            durum = "VAR"
                            tanimli+=1
                        else:
                            durum = "YOK"
                            tanimsiz+=1
                         
                temp_list_person_data.append(durum)
                list_persons_data.append(temp_list_person_data)
                    
                print(list_persons_data)
            
            print(list_persons_data)
            self.lbl_toplam_kayit.setText(str(toplam))
            self.lbl_tanimli.setText(str(tanimli))
            self.lbl_tanimsiz.setText(str(toplam-tanimli))
            
            self.listView_result.setColumnCount(len(list_persons_data[0]))
            self.listView_result.setRowCount(len(list_persons_data))
            self.listView_result.horizontalHeader().setStretchLastSection(True)
            #self.listView_result.resizeColumnsToContents()
            
            for i,row in enumerate(list_persons_data):
                for j,cell in enumerate(row):
                    self.listView_result.setItem(i,j, QTableWidgetItem(cell))
        else:
            self.showdialog("Warning","Yüz Algılanamadı","Belirtilen görüntü içerisinde yüz algılanamadığı için işleme devam edemessiniz.")
    
    def get_SCORE(self,img1,img2):
        print("checked----->",self.chk_SSIM.isChecked())
        if(self.chk_SSIM.isChecked()):
            score = self._ssim(img1,img2)
        if(self.chk_MSE.isChecked()):
            score = self._mse(img1, img2)
        if(self.chk_PSNR.isChecked()):
            score = self.psnr(img1,img2)
        """if(self.chk_all.isChecked()):
            score1 = self._ssim(img1,img2)
            score2 = self._mse(img1, img2)
            score2 = 1 - score2
            score3 = 100 / self.psnr(img1,img2)
            
            normalize ()
            score = (score1+score2)/2"""
    
        return score
    
    def Btn_Classes_Load(self):  
        table=self.Load_Excel()
        self.listView_class.setColumnCount(3)
        self.listView_class.setRowCount(len(table))
        
        for i,row in enumerate(table):
            for j,cell in enumerate(row):
                self.listView_class.setItem(i,j, QTableWidgetItem(str(cell)))
                #print(i,j,cell)
        
        self.listView_class.horizontalHeader().setStretchLastSection(True)
        self.listView_class.resizeColumnsToContents()
                
    def Load_Excel(self):
        table=[]          
        wb = load_workbook(filename = self.directory_recognition_classes+self.classes_index_file+".xlsx")
        sheet = wb.get_active_sheet()
        
        for i,row in enumerate(sheet.iter_rows()):  
            line=[]                                
            for j,cell in enumerate(row):                  
               line.append(cell.internal_value)
            table.append(line)
        
        return table
 
    def Btn_Image_Load(self):
        file,_ = QFileDialog.getOpenFileName(self, 'Open file', './',"Image files (*.png *.gif)")
        file_image = io.imread(file)
        cv2.imwrite(self._file_path,file_image)
        
        scene = self.show_image(r""+self._file_path)
        self.img_viewer.setScene(scene)
    
    def Btn_Camera(self):
        self.status_capture = False
        self.show_webcam(mirror=True)
    
    def Btn_Camera_Capture(self):
        self.status_capture=True
    
    #metodlar
    def show_image(self, img_path):
        self.pixmap = Qt.QPixmap()
        self.pixmap.load(img_path)
        self.pixmap = self.pixmap.scaled(self.img_viewer.size(), Qt.Qt.KeepAspectRatioByExpanding,transformMode=QtCore.Qt.SmoothTransformation)

        self.graphicsPixmapItem = Qt.QGraphicsPixmapItem(self.pixmap)

        self.graphicsScene = override_graphicsScene(self)
        self.graphicsScene.addItem(self.graphicsPixmapItem)
        
        self.img_viewer.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.img_viewer.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.img_person.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.img_person.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        
        return self.graphicsScene
    
    status_capture = False
    def show_webcam(self,mirror=False):
        cam = cv2.VideoCapture(0)
        while True:
            ret_val, img = cam.read()
            if mirror:
                if(self.status_capture):
                    img = cv2.flip(img, 1)
                    cv2.imwrite(self._file_path, img)
                    scene = self.show_image(self._file_path)
                    self.img_viewer.setScene(scene)
                    cv2.destroyAllWindows()
                    break
                else:
                    img = cv2.flip(img, 1)
            cv2.imshow('my webcam', img)
            if cv2.waitKey(1) == 27: 
                break  # esc to quit
        cv2.destroyAllWindows()

    def _ssim(self,img1,img2):
        img_1 = np.asarray(img1)#cv2.imread(img1)
        img_2 = np.asarray(img2)#cv2.imread(img2)
        return round(_SSIM(img_1,img_2),2)
        
    
    def _mse(self,img1,img2):
        img_1 = np.asarray(img1)#cv2.imread(img1)
        img_2 = np.asarray(img2)#cv2.imread(img2)
        
        e = np.sum((img_1.astype("float") - img_2.astype("float"))**2)
        e /= float(img_1.shape[0] * img_2.shape[1])
        r = round(e,2)
        print("MSE----->",str((r/10000)/2))
        if(r==0):
            return 1
        else:
            return 0
    
    def psnr(self,img1, img2):
        img1 = np.asarray(img1)#cv2.imread(img1)
        img2 = np.asarray(img2)#cv2.imread(img2)
        mse = np.mean( (img1 - img2) ** 2 )
        if mse == 0:
            return 1
        PIXEL_MAX = 255.0
        psnr = 20 * math.log10(PIXEL_MAX / math.sqrt(mse))
        print("PSNR------>",str(psnr/1000))
        return psnr/1000
    
    
    def rgb2gray(self,rgb):   
        r, g, b = rgb[:,:,0], rgb[:,:,1], rgb[:,:,2]
        gray = 0.2989 * r + 0.5870 * g + 0.1140 * b
    
        return gray
    
    def CLAHE(self,img):
        #img = color.rgb2gray(img)
        img = cv2.cvtColor(img,cv2.COLOR_BGR2LAB)
        img = cv2.split(img)
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
        img = clahe.apply(img[0])
        return img